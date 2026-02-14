import http.server
import socketserver
import json
import os
from datetime import datetime
try:
    import openpyxl
except ImportError:
    openpyxl = None

PORT = 3000
LOG_FILE = "logs.jsonl"
USERS_FILE = "users.json"
DATA_FILE = "data.xlsx"

class Handler(http.server.SimpleHTTPRequestHandler):
    def do_GET(self):
        if self.path == '/api/logs':
            self.send_response(200)
            self.send_header('Content-type', 'application/json')
            self.end_headers()
            
            logs = []
            if os.path.exists(LOG_FILE):
                with open(LOG_FILE, 'r', encoding='utf-8') as f:
                    for line in f:
                        if line.strip():
                            try:
                                logs.append(json.loads(line))
                            except:
                                pass
            
            # Return most recent first
            self.wfile.write(json.dumps(logs[::-1]).encode('utf-8'))
            
        elif self.path == '/api/users':
            self.send_response(200)
            self.send_header('Content-type', 'application/json')
            self.end_headers()
            
            users = {}
            if os.path.exists(USERS_FILE):
                try:
                    with open(USERS_FILE, 'r', encoding='utf-8') as f:
                        users = json.load(f)
                except:
                    pass
            self.wfile.write(json.dumps(users).encode('utf-8'))
            
        else:
            super().do_GET()

    def do_POST(self):
        content_length = int(self.headers['Content-Length'])
        post_data = self.rfile.read(content_length)
        
        if self.path == '/api/log':
            try:
                data = json.loads(post_data.decode('utf-8'))
                data['timestamp'] = datetime.now().isoformat()
                # Capture IP
                data['ip'] = self.client_address[0]
                
                with open(LOG_FILE, 'a', encoding='utf-8') as f:
                    f.write(json.dumps(data) + "\n")
                
                self.send_response(200)
                self.send_header('Content-type', 'application/json')
                self.end_headers()
                self.wfile.write(b'{"status":"ok"}')
            except Exception as e:
                print(f"Error logging: {e}")
                self.send_response(500)
                self.end_headers()
                
        elif self.path == '/api/users':
            try:
                req = json.loads(post_data.decode('utf-8'))
                ip = req.get('ip')
                name = req.get('name')
                
                if ip and name:
                    users = {}
                    if os.path.exists(USERS_FILE):
                        try:
                            with open(USERS_FILE, 'r', encoding='utf-8') as f:
                                users = json.load(f)
                        except:
                            pass
                    
                    users[ip] = name
                    
                    with open(USERS_FILE, 'w', encoding='utf-8') as f:
                        json.dump(users, f, indent=2)
                        
                    self.send_response(200)
                    self.wfile.write(b'{"status":"ok"}')
                else:
                    self.send_error(400)
            except Exception as e:
                print(f"Error saving user: {e}")
                self.send_response(500)
                self.end_headers()

        elif self.path == '/api/update-intersection':
            if not openpyxl:
                self.send_error(500, "openpyxl not installed")
                return

            try:
                req = json.loads(post_data.decode('utf-8'))
                route_id = req.get('route')
                intersection = req.get('intersection')
                updates = req.get('updates', {}) # {name, coordinates, plans}

                if not os.path.exists(DATA_FILE):
                     self.send_error(404, "Data file not found")
                     return

                wb = openpyxl.load_workbook(DATA_FILE)
                # Assuming data is in the first sheet or specific sheets.
                # Based on js/data.js, we merge Sheet1 and Sheet2.
                # We need to find where this record lives.
                # Simple strategy: Search both sheets for the matching Route and Intersection
                
                found = False
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    
                    # Find headers to identify columns
                    headers = {}
                    for cell in ws[1]:
                        if cell.value:
                            headers[str(cell.value).lower()] = cell.column_letter

                    # We need at least 'route' and 'streetsort' (or 'street') to identify
                    route_col = headers.get('route')
                    street_col = headers.get('streetsort') or headers.get('street') or headers.get('intersection')
                    
                    if not route_col or not street_col:
                        continue

                    # Iterate rows
                    for row in ws.iter_rows(min_row=2):
                        # Get values for identification
                        # openpyxl uses 1-based indexing for columns if accessing by index, but we have column letters
                        r_val = ws[f"{route_col}{row[0].row}"].value
                        s_val = ws[f"{street_col}{row[0].row}"].value
                        
                        # Match logic (relaxed comparison)
                        if str(r_val) == str(route_id) and str(s_val) == str(intersection):
                            found = True
                            r_idx = row[0].row
                            
                            # Apply updates
                            # 1. Name (Intersection/Street)
                            if 'name' in updates:
                                ws[f"{street_col}{r_idx}"] = updates['name']
                            
                            # 2. Coordinates
                            if 'coordinates' in updates:
                                coord_col = headers.get('coordinates')
                                if coord_col:
                                    ws[f"{coord_col}{r_idx}"] = updates['coordinates']
                            
                            # 3. Plans (6 car, etc)
                            # These might be in different columns depending on the sheet
                            if 'plans' in updates:
                                for plan_key, plan_val in updates['plans'].items():
                                    # plan_key expected like "6 car", "4 car"
                                    # Normalize key to match header
                                    p_col = headers.get(plan_key.lower())
                                    if p_col:
                                        ws[f"{p_col}{r_idx}"] = plan_val
                            
                            break # modified the row, stop searching this sheet
                    
                    if found:
                        break # Stop searching other sheets if found

                if found:
                    wb.save(DATA_FILE)
                    self.send_response(200)
                    self.wfile.write(b'{"status":"ok"}')
                else:
                    self.send_error(404, "Intersection not found")

            except Exception as e:
                print(f"Error updating excel: {e}")
                self.send_response(500, f"Error: {str(e)}")
                self.end_headers()

        else:
            self.send_error(404)

print(f"Server started at http://localhost:{PORT}")
with socketserver.TCPServer(("", PORT), Handler) as httpd:
    httpd.serve_forever()
